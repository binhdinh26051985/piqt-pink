import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.dates as mdates
from datetime import datetime

# Read work hours data
workhr = pd.read_excel(r"d:\IE\EFFCAL\Attendance.xlsx", sheet_name='Attend')

# Convert Date column to datetime if it's not already
if not pd.api.types.is_datetime64_any_dtype(workhr['Date']):
    workhr['Date'] = pd.to_datetime(workhr['Date'])

workhr['WorkingTT(hrs)'] = workhr['Attendance'] * workhr['Working Hrs']
workhr['SKU'] = workhr['Line'].astype(str) + '_' + workhr['Date'].dt.strftime('%Y-%m-%d')
workhrs = workhr.groupby(['SKU'])['WorkingTT(hrs)'].sum().reset_index()
workhrs1 = workhrs[['SKU', 'WorkingTT(hrs)']]

# Read SMV data
smv = pd.read_excel(r"d:\IE\EFFCAL\Attendance.xlsx", sheet_name='SMV')    
smv = smv.drop_duplicates(subset=['Style'])

# Read output data
ouputhrs = pd.read_excel(r"d:\IE\EFFCAL\Attendance.xlsx", sheet_name='Output')

# Convert Date column to datetime if it's not already
if not pd.api.types.is_datetime64_any_dtype(ouputhrs['Date']):
    ouputhrs['Date'] = pd.to_datetime(ouputhrs['Date'])

ouputhrs = ouputhrs.merge(smv, on='Style', how='left')
ouputhrs['SmvTT(hrs)'] = ouputhrs['Output'] * ouputhrs['SMV'] / 60
ouputhrs['SKU'] = ouputhrs['Line'].astype(str) + '_' + ouputhrs['Date'].dt.strftime('%Y-%m-%d')

ouputhrs2 = ouputhrs.groupby(['SKU'])['SmvTT(hrs)'].sum().reset_index()

# Merge work hours and SMV hours
efficiency = workhrs1.merge(ouputhrs2, on='SKU', how='left')

# Handle NaN values in SmvTT(hrs) - fill with 0
efficiency['SmvTT(hrs)'] = efficiency['SmvTT(hrs)'].fillna(0)

# Calculate efficiency with error handling
efficiency['Eff%'] = np.where(
    efficiency['WorkingTT(hrs)'] > 0,
    (efficiency['SmvTT(hrs)'] / efficiency['WorkingTT(hrs)']) * 100,
    0
)

# Extract Date and Line from SKU
efficiency['Date'] = efficiency['SKU'].str.extract(r'_(\d{4}-\d{2}-\d{2})')
efficiency['Line'] = efficiency['SKU'].str.split('_').str[0]

# Convert Date to datetime format
efficiency['Date'] = pd.to_datetime(efficiency['Date'])

# Add month column
efficiency['Month'] = efficiency['Date'].dt.to_period('M')

# IMPORTANT: Sort by Date for chronological plotting
efficiency = efficiency.sort_values('Date')

# Group by Month and Line
monthly_summary = efficiency.groupby(['Month', 'Line']).agg(
    WorkingTT_total=('WorkingTT(hrs)', 'sum'),
    SmvTT_total=('SmvTT(hrs)', 'sum'),
).reset_index()

# Calculate overall efficiency from totals with error handling
monthly_summary['Efficiency_from_total'] = np.where(
    monthly_summary['WorkingTT_total'] > 0,
    (monthly_summary['SmvTT_total'] / monthly_summary['WorkingTT_total']) * 100,
    0
)

# Format for better readability
monthly_summary['Month'] = monthly_summary['Month'].astype(str)
monthly_summary = monthly_summary.round(2)

# Sort by Month and Line
monthly_summary_line = monthly_summary.sort_values(['Month', 'Line'])

# Overall monthly summary
monthly_eff = efficiency.groupby(['Month']).agg(
    WorkingTT_total=('WorkingTT(hrs)', 'sum'),
    SmvTT_total=('SmvTT(hrs)', 'sum'),
).reset_index()

# Calculate efficiency with error handling
monthly_eff['Eff%'] = np.where(
    monthly_eff['WorkingTT_total'] > 0,
    (monthly_eff['SmvTT_total'] / monthly_eff['WorkingTT_total']) * 100,
    0
)

# Round the efficiency values
monthly_eff['Eff%'] = monthly_eff['Eff%'].round(2)
monthly_eff = monthly_eff.sort_values('Month')
monthly_eff['Month_str'] = monthly_eff['Month'].astype(str)

# Create summary statistics
summary_stats = efficiency.groupby('Line')['Eff%'].agg(['mean', 'median', 'min', 'max', 'std']).round(2)

# Create Monthly Average Efficiency by Line
monthly_avg = efficiency.groupby(['Month', 'Line'])['Eff%'].mean().reset_index()
monthly_avg['Month'] = monthly_avg['Month'].astype(str)
monthly_avg = monthly_avg.sort_values(['Month', 'Line'])

# ===========================================
# CREATE CHARTS
# ===========================================

# Chart 1: Monthly Overall Efficiency
plt.style.use('seaborn-v0_8-darkgrid')
plt.figure(figsize=(12, 6))
plt.plot(monthly_eff['Month_str'], monthly_eff['Eff%'], 
         marker='o', linewidth=2.5, markersize=8, color='steelblue')
plt.title('Monthly Overall Efficiency', fontsize=16, fontweight='bold', pad=20)
plt.xlabel('Month', fontsize=12)
plt.ylabel('Efficiency (%)', fontsize=12)
plt.xticks(rotation=45, ha='right')
plt.grid(True, alpha=0.3)
for x, y in zip(monthly_eff['Month_str'], monthly_eff['Eff%']):
    plt.text(x, y + 0.5, f'{y:.1f}%', ha='center', va='bottom', fontsize=9)
plt.ylim(bottom=0)
plt.tight_layout()
plt.savefig('chart1_monthly_overall.png', dpi=80, bbox_inches='tight')
plt.close()

# Chart 2: Daily Efficiency by Line
plt.figure(figsize=(16, 8))
lines = efficiency['Line'].unique()
colors = plt.cm.Set2(np.linspace(0, 1, len(lines)))

for line, color in zip(sorted(lines), colors):
    line_data = efficiency[efficiency['Line'] == line]
    if not line_data.empty:
        plt.plot(line_data['Date'], line_data['Eff%'], 
                 marker='o', linewidth=2, markersize=4, 
                 label=f'{line}', color=color, alpha=0.8)

plt.title('Daily Efficiency Trend by Production Line', fontsize=16, fontweight='bold', pad=20)
plt.xlabel('Date', fontsize=12)
plt.ylabel('Efficiency (%)', fontsize=12)
plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
plt.gcf().autofmt_xdate(rotation=45)
plt.grid(True, alpha=0.3)

# Only add legend if we have labels
handles, labels = plt.gca().get_legend_handles_labels()
if labels:
    plt.legend(title='Production Line', bbox_to_anchor=(1.05, 1), loc='upper left')

plt.axhline(y=100, color='red', linestyle='--', alpha=0.5, label='100% Target')
max_eff = efficiency['Eff%'].max()
plt.ylim(0, max(120, max_eff * 1.1))
plt.tight_layout()
plt.savefig('chart2_daily_by_line.png', dpi=150, bbox_inches='tight')
plt.close()

# Chart 3: Monthly Average Efficiency by Line (with specified colors)
plt.figure(figsize=(14, 7))

# Create a color mapping dictionary for each line
line_colors = {
    '1': 'red',
    '2': 'purple', 
    '3': 'blue',
    '4': 'black',
    '5': 'pink',
    '6': 'yellow'
}

# Define line numbers in order
line_numbers = ['1', '2', '3', '4', '5', '6']

# Plot lines in specified order with specified colors
has_labels = False
for line_num in line_numbers:
    line_data = monthly_avg[monthly_avg['Line'] == line_num]
    if not line_data.empty:
        color = line_colors.get(line_num, plt.cm.Set2(0))
        plt.plot(line_data['Month'], line_data['Eff%'], 
                 marker='s', linewidth=2.5, markersize=6, 
                 label=f'Line {line_num}', color=color)
        has_labels = True

plt.title('Monthly Average Efficiency by Production Line', fontsize=16, fontweight='bold', pad=20)
plt.xlabel('Month', fontsize=12)
plt.ylabel('Average Efficiency (%)', fontsize=12)
plt.xticks(rotation=45)
plt.axhline(y=100, color='red', linestyle='--', alpha=0.5)

# Only add legend if we have labels
if has_labels:
    plt.legend(title='Production Line', bbox_to_anchor=(1.05, 1), loc='upper left')

plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('chart3_monthly_avg_by_line.png', dpi=150, bbox_inches='tight')
plt.close()

# Chart 4: Line-wise Efficiency Distribution (Box Plot)
plt.figure(figsize=(12, 6))
efficiency_sorted = efficiency.copy()
# Create a custom order for lines
line_order = sorted(efficiency['Line'].unique())
efficiency_sorted['Line'] = pd.Categorical(efficiency_sorted['Line'], categories=line_order, ordered=True)

# Fix for Matplotlib 3.9+ deprecation warning
plt.boxplot([efficiency_sorted[efficiency_sorted['Line'] == line]['Eff%'] for line in line_order],
            tick_labels=line_order)  # Changed from 'labels' to 'tick_labels'
plt.title('Efficiency Distribution by Production Line', fontsize=16, fontweight='bold', pad=20)
plt.xlabel('Production Line', fontsize=12)
plt.ylabel('Efficiency (%)', fontsize=12)
plt.grid(True, alpha=0.3)
plt.axhline(y=100, color='red', linestyle='--', alpha=0.5)
plt.tight_layout()
plt.savefig('chart4_efficiency_distribution.png', dpi=150, bbox_inches='tight')
plt.close()

# ===========================================
# CREATE INDIVIDUAL DAILY EFFICIENCY CHARTS FOR EACH LINE (6 CHARTS)
# ===========================================

# Get unique lines
lines = sorted(efficiency['Line'].unique())

# Colors for each line (same as Chart 3 colors)
line_colors = {
    '1': 'red',
    '2': 'purple', 
    '3': 'blue',
    '4': 'black',
    '5': 'pink',
    '6': 'yellow'
}

# Create individual charts for each line
for line in lines:
    plt.figure(figsize=(14, 7))
    
    # Filter data for this line
    line_data = efficiency[efficiency['Line'] == line].copy()
    line_data = line_data.sort_values('Date')  # Ensure chronological order
    
    if not line_data.empty:
        # Get color for this line
        color = line_colors.get(str(line), plt.cm.Set2(0))
        
        # Plot the line's daily efficiency
        plt.plot(line_data['Date'], line_data['Eff%'], 
                 marker='o', linewidth=2.5, markersize=6, 
                 label=f'Line {line}', color=color, alpha=0.9)
        
        plt.title(f'Daily Efficiency Trend - Line {line}', fontsize=16, fontweight='bold', pad=20)
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Efficiency (%)', fontsize=12)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gcf().autofmt_xdate(rotation=45)
        
        # Calculate statistics for this line
        avg_eff = line_data['Eff%'].mean()
        max_eff = line_data['Eff%'].max()
        min_eff = line_data['Eff%'].min()
        
        # Add statistics as text
        stats_text = f'Average: {avg_eff:.1f}%\nMax: {max_eff:.1f}%\nMin: {min_eff:.1f}%'
        plt.text(0.02, 0.98, stats_text, transform=plt.gca().transAxes,
                 fontsize=11, verticalalignment='top',
                 bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        # Add horizontal lines and targets
        plt.axhline(y=100, color='red', linestyle='--', alpha=0.5, label='100% Target')
        plt.axhline(y=avg_eff, color='green', linestyle='--', alpha=0.5, label=f'Avg: {avg_eff:.1f}%')
        
        plt.grid(True, alpha=0.3)
        plt.legend(loc='upper left')
        
        # Set y-axis limits
        y_max = max(120, max_eff * 1.1)
        plt.ylim(0, y_max)
        
        plt.tight_layout()
        
        # Save each chart with line-specific name
        plt.savefig(f'chart_line_{line}_daily.png', dpi=150, bbox_inches='tight')
        plt.close()
    else:
        print(f"No data found for Line {line}")

# ===========================================
# CREATE EXCEL REPORT WITH IMAGES
# ===========================================

# First, save all data to Excel
with pd.ExcelWriter('Efficiency_Report.xlsx', engine='openpyxl') as writer:
    # Write data sheets
    efficiency.to_excel(writer, sheet_name='Daily_Data', index=False)
    monthly_summary_line.to_excel(writer, sheet_name='Monthly_by_Line', index=False)
    monthly_eff.to_excel(writer, sheet_name='Monthly_Overall', index=False)
    summary_stats.to_excel(writer, sheet_name='Summary_Statistics')
    monthly_avg.to_excel(writer, sheet_name='Monthly_Avg_by_Line', index=False)
    
    # Create a summary dashboard sheet
    summary_df = pd.DataFrame({
        'Metric': ['Total Days Analyzed', 'Number of Lines', 'Date Range', 
                   'Overall Avg Efficiency', 'Highest Efficiency', 'Lowest Efficiency'],
        'Value': [efficiency['Date'].nunique(), 
                  efficiency['Line'].nunique(),
                  f"{efficiency['Date'].min().date()} to {efficiency['Date'].max().date()}",
                  f"{efficiency['Eff%'].mean():.1f}%",
                  f"{efficiency['Eff%'].max():.1f}%",
                  f"{efficiency['Eff%'].min():.1f}%"]
    })
    summary_df.to_excel(writer, sheet_name='Dashboard', index=False)

# Now add images to the Excel file
wb = load_workbook('Efficiency_Report.xlsx')

# Add Chart 1 to Monthly_Overall sheet
ws = wb['Monthly_Overall']
img1 = Image('chart1_monthly_overall.png')
# Position after data (find last row)
max_row = ws.max_row
img1.anchor = f'E{max_row + 3}'  # Place 3 rows below the data
ws.add_image(img1)

# Add Chart 2 to Daily_Data sheet
ws = wb['Daily_Data']
img2 = Image('chart2_daily_by_line.png')
max_row = ws.max_row
img2.anchor = f'H{3}'  # Place at column H, row 3
ws.add_image(img2)

# Add Chart 3 to Monthly_Avg_by_Line sheet
ws = wb['Monthly_Avg_by_Line']
img3 = Image('chart3_monthly_avg_by_line.png')
max_row = ws.max_row
img3.anchor = f'G{3}'  # Place at column G, row 3
ws.add_image(img3)

# Add Chart 4 to Summary_Statistics sheet
ws = wb['Summary_Statistics']
img4 = Image('chart4_efficiency_distribution.png')
max_row = ws.max_row
img4.anchor = f'G{3}'  # Place at column G, row 3
ws.add_image(img4)

# Add individual line charts to a new sheet
if 'Line_Charts' not in wb.sheetnames:
    wb.create_sheet('Line_Charts')

ws_line_charts = wb['Line_Charts']
ws_line_charts.title = 'Line_Charts'

# Add each individual line chart to the sheet
current_row = 1
for i, line in enumerate(lines):
    chart_file = f'chart_line_{line}_daily.png'
    if os.path.exists(chart_file):
        img = Image(chart_file)
        
        # Position charts in 2x3 grid
        if i < 3:  # First row
            col_offset = i * 15  # 15 columns between charts
            row_offset = 0
        else:  # Second row
            col_offset = (i - 3) * 15
            row_offset = 25  # 25 rows between rows
            
        # Convert column number to letter (fix for columns beyond Z)
        if col_offset < 26:
            col_letter = chr(65 + col_offset)
        elif col_offset < 52:
            col_letter = 'A' + chr(65 + col_offset - 26)
        else:
            col_letter = 'BA'  # Default to BA if beyond
            
        # Ensure row number is at least 1
        row_number = current_row + row_offset
        if row_number < 1:
            row_number = 1
            
        img.anchor = f'{col_letter}{row_number}'
        ws_line_charts.add_image(img)
        
        # Add line name above chart (ensure row is at least 1)
        title_row = row_number - 1
        if title_row < 1:
            title_row = 1
        ws_line_charts[f'{col_letter}{title_row}'] = f'Line {line} Daily Efficiency'

# Add a note to Dashboard sheet
ws = wb['Dashboard']
ws['D10'] = 'Report Generated On:'
ws['E10'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# Adjust column widths for better readability
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook with all images
wb.save('Efficiency_Report.xlsx')

# Clean up temporary image files
image_files = ['chart1_monthly_overall.png', 'chart2_daily_by_line.png', 
               'chart3_monthly_avg_by_line.png', 'chart4_efficiency_distribution.png']

# Add individual line chart files to cleanup list
line_chart_files = [f'chart_line_{line}_daily.png' for line in lines]
image_files.extend(line_chart_files)

for img_file in image_files:
    if os.path.exists(img_file):
        os.remove(img_file)

